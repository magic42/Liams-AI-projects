#!/usr/bin/env python3
"""
Universal Website Scraper Agent

A configurable scraper that can be called with simple parameters to crawl
any website and extract images from categories, products, or the entire site.

Usage:
    source ~/.scrapy-crawler-venv/bin/activate

    # Scrape category pages
    python scraper.py --url "www.example.com" --type category

    # Scrape product pages
    python scraper.py --url "www.example.com" --type product

    # Scrape entire site
    python scraper.py --url "www.example.com" --type all

    # With options
    python scraper.py --url "www.example.com" --type category --max-pages 50 --delay 0.5

Output files are named: {domain}-{type}.csv and {domain}-{type}_unique.csv
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

from scrapy import Spider, signals
from scrapy.crawler import CrawlerProcess
from scrapy.linkextractors import LinkExtractor
from scrapy.spidermiddlewares.httperror import HttpError
from twisted.internet.error import DNSLookupError, TCPTimedOutError, TimeoutError


# =============================================================================
# SCRAPE TYPE CONFIGURATIONS
# =============================================================================

SCRAPE_CONFIGS = {
    "category": {
        "url_patterns": [
            r'/category/',
            r'/categories/',
            r'/cat/',
            r'/c/',
            r'/collections/',
            r'/shop/',
            r'/browse/',
        ],
        "description": "Category/collection pages",
    },
    "product": {
        "url_patterns": [
            r'/product/',
            r'/products/',
            r'/p/',
            r'/item/',
            r'/items/',
            r'/dp/',  # Amazon style
            r'/pd/',
        ],
        "description": "Product detail pages",
    },
    "blog": {
        "url_patterns": [
            r'/blog/',
            r'/news/',
            r'/articles/',
            r'/posts/',
            r'/journal/',
        ],
        "description": "Blog/news pages",
    },
    "all": {
        "url_patterns": [],  # Empty = follow all internal links
        "description": "Entire website",
    },
    "fullmonty": {
        "url_patterns": [
            # Category patterns
            r'/category/',
            r'/categories/',
            r'/cat/',
            r'/c/',
            r'/collections/',
            r'/shop/',
            r'/browse/',
            # Product patterns
            r'/product/',
            r'/products/',
            r'/p/',
            r'/item/',
            r'/items/',
            r'/dp/',
            r'/pd/',
        ],
        "description": "Full Monty - categories + products with full data extraction",
    },
}

# Common URL patterns to always exclude
DEFAULT_DENY_PATTERNS = [
    r'/cart/',
    r'/checkout/',
    r'/account/',
    r'/login/',
    r'/register/',
    r'/my-account/',
    r'/admin/',
    r'/wp-admin/',
    r'/wp-login/',
    r'\?add-to-cart=',
    r'\?remove_item=',
    r'/wishlist/',
    r'/compare/',
]

# Image patterns to exclude (icons, placeholders, etc.)
IMAGE_EXCLUDE_PATTERNS = [
    r'placeholder',
    r'loading',
    r'spinner',
    r'icon',
    r'pixel',
    r'tracking',
    r'spacer',
    r'blank',
    r'1x1',
    r'transparent',
    r'/wp-includes/',
    r'/wp-content/plugins/',
    r'gravatar\.com',
]

# Compiled patterns for page classification (used by Full Monty)
CATEGORY_PATTERNS = [re.compile(p, re.IGNORECASE) for p in SCRAPE_CONFIGS["category"]["url_patterns"]]
PRODUCT_PATTERNS = [re.compile(p, re.IGNORECASE) for p in SCRAPE_CONFIGS["product"]["url_patterns"]]


def classify_page(url, known_product_urls=None, known_category_urls=None):
    """Classify a URL as 'product', 'category', or 'other'.

    If known_product_urls or known_category_urls sets are provided,
    classification is done by lookup first (for sites with non-standard
    URL patterns). Falls back to regex pattern matching.
    """
    # Priority 1: Check against explicitly known URL sets
    if known_product_urls and url in known_product_urls:
        return "product"
    if known_category_urls and url in known_category_urls:
        return "category"

    # Priority 2: Regex pattern matching on URL path
    path = urlparse(url).path
    for pattern in PRODUCT_PATTERNS:
        if pattern.search(path):
            return "product"
    for pattern in CATEGORY_PATTERNS:
        if pattern.search(path):
            return "category"
    return "other"


# =============================================================================
# SPIDER IMPLEMENTATION
# =============================================================================


class ScraperSpider(Spider):
    """Universal scraper spider with configurable URL patterns."""

    name = "scraper"

    def __init__(self, config, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.config = config
        self.results = []
        self.errors = []
        self.pages_crawled = 0
        self.all_images = {}

        # Known URL sets for page classification (used by Full Monty)
        self.known_product_urls = set(config.get("known_product_urls", []))
        self.known_category_urls = set(config.get("known_category_urls", []))

        # Compile image exclusion patterns
        self.image_exclude_patterns = [
            re.compile(p, re.IGNORECASE) for p in IMAGE_EXCLUDE_PATTERNS
        ]

        # Set up allowed domains
        domain = config["domain"]
        self.allowed_domains = [domain, domain.replace("www.", "")]

        # Set start URLs
        self.start_urls = [f"https://{domain}/"]

        # Add extra seed URLs from file if provided
        extra_urls = config.get("seed_urls", [])
        if extra_urls:
            for url in extra_urls:
                url = url.strip()
                if url and url not in self.start_urls:
                    self.start_urls.append(url)

        # Configure link extractor based on scrape type
        scrape_type = config["scrape_type"]
        url_patterns = SCRAPE_CONFIGS[scrape_type]["url_patterns"]

        self.link_extractor = LinkExtractor(
            allow_domains=self.allowed_domains,
            allow=url_patterns if url_patterns else (),
            deny=DEFAULT_DENY_PATTERNS,
            deny_extensions=[
                "jpg", "jpeg", "png", "gif", "svg", "webp", "ico", "bmp",
                "pdf", "doc", "docx", "xls", "xlsx", "ppt", "pptx",
                "mp3", "mp4", "avi", "mov", "wmv", "flv",
                "zip", "rar", "tar", "gz", "7z",
                "css", "js", "woff", "woff2", "ttf", "eot",
            ],
        )

    @classmethod
    def create_settings(cls, config):
        """Create Scrapy settings from config."""
        return {
            "DOWNLOAD_DELAY": config.get("delay", 1.0),
            "CONCURRENT_REQUESTS": config.get("concurrent", 2),
            "CONCURRENT_REQUESTS_PER_DOMAIN": config.get("concurrent", 2),
            "ROBOTSTXT_OBEY": True,
            "USER_AGENT": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36 "
                "(Magic42Scraper/1.0)"
            ),
            "LOG_LEVEL": config.get("log_level", "INFO"),
            "DOWNLOAD_TIMEOUT": 30,
            "RETRY_TIMES": 2,
            "RETRY_HTTP_CODES": [500, 502, 503, 504, 408, 429],
            "HTTPERROR_ALLOW_ALL": True,
        }

    def parse(self, response):
        """Parse a page and extract images."""
        max_pages = self.config.get("max_pages", 0)
        if max_pages > 0 and self.pages_crawled >= max_pages:
            self.crawler.engine.close_spider(self, "max_pages_reached")
            return

        if response.status != 200:
            self._handle_http_error(response)
            if response.status >= 400:
                return

        self.pages_crawled += 1

        # Extract images
        images = self._extract_images(response)

        # Full Monty: extract all data (SEO, product, content metrics)
        if self.config["scrape_type"] == "fullmonty":
            meta_data = self._extract_meta_data(response)
            content_metrics = self._extract_content_metrics(response)
            structured_data = self._extract_structured_data(response)
            page_type = classify_page(
                response.url,
                known_product_urls=self.known_product_urls,
                known_category_urls=self.known_category_urls,
            )

            result = {
                "page_url": response.url,
                "page_type": page_type,
                # SEO Data
                "meta_title": meta_data["meta_title"],
                "meta_description": meta_data["meta_description"],
                "h1": meta_data["h1"],
                "canonical_url": meta_data["canonical_url"],
                "og_image": meta_data["og_image"],
                "og_title": meta_data["og_title"],
                "og_description": meta_data["og_description"],
                # Content Metrics
                "word_count": content_metrics["word_count"],
                "total_image_count": content_metrics["total_image_count"],
                "internal_link_count": content_metrics["internal_link_count"],
                "external_link_count": content_metrics["external_link_count"],
                # Product Data
                "product_name": structured_data["product_name"],
                "product_price": structured_data["product_price"],
                "product_currency": structured_data["product_currency"],
                "product_sku": structured_data["product_sku"],
                "product_brand": structured_data["product_brand"],
                "product_availability": structured_data["product_availability"],
                "product_description": structured_data["product_description"],
                # Schema
                "has_schema_markup": structured_data["has_schema_markup"],
                "schema_types": structured_data["schema_types"],
                # Images
                "image_count": len(images),
                "images": json.dumps(images),
            }
            self.results.append(result)

            self.logger.info(
                f"Crawled [{page_type}]: {response.url} - "
                f"{len(images)} images, price={structured_data['product_price'] or 'N/A'}"
            )
        else:
            # Standard mode: only store results when images found
            page_title = (
                response.xpath("//title/text()").get() or
                response.xpath("//meta[@property='og:title']/@content").get() or
                ""
            ).strip()

            if images:
                result = {
                    "page_url": response.url,
                    "page_title": page_title,
                    "image_count": len(images),
                    "images": json.dumps(images),
                }
                self.results.append(result)

                self.logger.info(
                    f"Crawled: {response.url} - Found {len(images)} images"
                )
            else:
                self.logger.info(f"Crawled: {response.url} - No images found")

        # Track unique images (all modes)
        for img in images:
            img_url = img["src"]
            if img_url not in self.all_images:
                self.all_images[img_url] = {
                    "src": img_url,
                    "alt": img.get("alt", ""),
                    "found_on": [response.url],
                }
            else:
                self.all_images[img_url]["found_on"].append(response.url)

        # Follow links
        for link in self.link_extractor.extract_links(response):
            if max_pages > 0 and self.pages_crawled >= max_pages:
                break
            yield response.follow(
                link.url,
                callback=self.parse,
                errback=self._handle_error,
            )

    def _extract_images(self, response):
        """Extract and filter images from the page."""
        images = []
        seen_urls = set()
        min_width = self.config.get("min_image_width", 50)
        min_height = self.config.get("min_image_height", 50)

        for img in response.xpath("//img"):
            src = (
                img.xpath("@src").get() or
                img.xpath("@data-src").get() or
                img.xpath("@data-lazy-src").get() or
                img.xpath("@data-original").get() or
                ""
            )

            if not src or src.startswith("data:"):
                continue

            src = urljoin(response.url, src)

            if src in seen_urls:
                continue
            seen_urls.add(src)

            if self._should_exclude_image(src):
                continue

            alt = img.xpath("@alt").get() or ""
            width = img.xpath("@width").get()
            height = img.xpath("@height").get()

            try:
                w = int(width) if width else 0
                h = int(height) if height else 0
                if min_width > 0 and w > 0 and w < min_width:
                    continue
                if min_height > 0 and h > 0 and h < min_height:
                    continue
            except (ValueError, TypeError):
                pass

            images.append({
                "src": src,
                "alt": alt.strip(),
                "width": width,
                "height": height,
            })

        # Check for background images
        for element in response.xpath("//*[@style]"):
            style = element.xpath("@style").get() or ""
            bg_match = re.search(r'url\(["\']?([^"\')\s]+)["\']?\)', style)
            if bg_match:
                src = urljoin(response.url, bg_match.group(1))
                if src not in seen_urls and not self._should_exclude_image(src):
                    seen_urls.add(src)
                    images.append({
                        "src": src,
                        "alt": "(background image)",
                        "width": None,
                        "height": None,
                    })

        return images

    def _should_exclude_image(self, url):
        """Check if image URL matches exclusion patterns."""
        return any(p.search(url) for p in self.image_exclude_patterns)

    def _extract_meta_data(self, response):
        """Extract SEO and meta data from the page."""
        return {
            "meta_title": (response.xpath("//title/text()").get() or "").strip(),
            "meta_description": (
                response.xpath("//meta[@name='description']/@content").get() or ""
            ).strip(),
            "h1": (response.xpath("//h1//text()").get() or "").strip(),
            "canonical_url": (
                response.xpath("//link[@rel='canonical']/@href").get() or ""
            ).strip(),
            "og_image": (
                response.xpath("//meta[@property='og:image']/@content").get() or ""
            ).strip(),
            "og_title": (
                response.xpath("//meta[@property='og:title']/@content").get() or ""
            ).strip(),
            "og_description": (
                response.xpath("//meta[@property='og:description']/@content").get() or ""
            ).strip(),
        }

    def _extract_content_metrics(self, response):
        """Extract content metrics from the page."""
        body_text = " ".join(response.xpath("//body//text()").getall())
        body_text = re.sub(r'\s+', ' ', body_text).strip()
        word_count = len(body_text.split()) if body_text else 0

        total_images = len(response.xpath("//img").getall())

        all_links = response.xpath("//a/@href").getall()
        domain = self.config["domain"]
        internal_links = 0
        external_links = 0
        for href in all_links:
            if not href or href.startswith("#") or href.startswith("javascript:"):
                continue
            full_url = urljoin(response.url, href)
            parsed = urlparse(full_url)
            if parsed.netloc and domain not in parsed.netloc:
                external_links += 1
            else:
                internal_links += 1

        return {
            "word_count": word_count,
            "total_image_count": total_images,
            "internal_link_count": internal_links,
            "external_link_count": external_links,
        }

    def _extract_structured_data(self, response):
        """Extract JSON-LD structured data from the page."""
        result = {
            "has_schema_markup": False,
            "schema_types": "",
            "product_name": "",
            "product_price": "",
            "product_currency": "",
            "product_sku": "",
            "product_brand": "",
            "product_availability": "",
            "product_description": "",
        }

        json_ld_scripts = response.xpath(
            '//script[@type="application/ld+json"]/text()'
        ).getall()

        schema_types = []
        for script_text in json_ld_scripts:
            try:
                data = json.loads(script_text)
                items = data if isinstance(data, list) else [data]
                for item in items:
                    self._process_schema_item(item, result, schema_types)
                    # Handle @graph pattern (common in Yoast/WooCommerce)
                    if "@graph" in item:
                        for graph_item in item["@graph"]:
                            self._process_schema_item(graph_item, result, schema_types)
            except (json.JSONDecodeError, TypeError, AttributeError):
                continue

        result["schema_types"] = ", ".join(schema_types)

        # Fallback: microdata/itemprop attributes
        if not result["product_name"]:
            result["product_name"] = (
                response.xpath('//*[@itemprop="name"]/text()').get() or ""
            ).strip()
        if not result["product_price"]:
            result["product_price"] = (
                response.xpath('//*[@itemprop="price"]/@content').get() or
                response.xpath('//*[@itemprop="price"]/text()').get() or ""
            ).strip()
        if not result["product_sku"]:
            result["product_sku"] = (
                response.xpath('//*[@itemprop="sku"]/@content').get() or
                response.xpath('//*[@itemprop="sku"]/text()').get() or ""
            ).strip()
        if not result["product_brand"]:
            result["product_brand"] = (
                response.xpath('//*[@itemprop="brand"]//text()').get() or ""
            ).strip()

        # Fallback: common CSS class patterns for price
        if not result["product_price"]:
            price_selectors = [
                '.price ::text', '.product-price ::text',
                '.current-price ::text', '[data-price] ::text',
                '.woocommerce-Price-amount ::text',
            ]
            for selector in price_selectors:
                price_text = response.css(selector).get()
                if price_text:
                    price_match = re.search(r'[\d,.]+', price_text.strip())
                    if price_match:
                        result["product_price"] = price_match.group(0)
                        break

        return result

    def _process_schema_item(self, item, result, schema_types):
        """Process a single schema.org item from JSON-LD."""
        if not isinstance(item, dict) or "@type" not in item:
            return

        schema_types.append(item["@type"])
        result["has_schema_markup"] = True

        if item["@type"] == "Product":
            result["product_name"] = item.get("name", "") or result["product_name"]
            result["product_description"] = item.get("description", "") or result["product_description"]
            result["product_sku"] = item.get("sku", "") or result["product_sku"]

            brand = item.get("brand")
            if isinstance(brand, dict):
                result["product_brand"] = brand.get("name", "") or result["product_brand"]
            elif brand:
                result["product_brand"] = str(brand) or result["product_brand"]

            offers = item.get("offers", {})
            if isinstance(offers, list):
                offers = offers[0] if offers else {}
            if isinstance(offers, dict):
                result["product_price"] = str(offers.get("price", "")) or result["product_price"]
                result["product_currency"] = offers.get("priceCurrency", "") or result["product_currency"]
                availability = offers.get("availability", "")
                result["product_availability"] = (
                    availability.replace("https://schema.org/", "").replace("http://schema.org/", "")
                    or result["product_availability"]
                )

    def _handle_http_error(self, response):
        """Handle HTTP errors."""
        self.errors.append({
            "url": response.url,
            "status": response.status,
            "error_type": "http_error",
            "timestamp": datetime.now().isoformat(),
        })
        if response.status == 429:
            self.logger.error(f"Rate limited (429): {response.url}")
        elif response.status >= 500:
            self.logger.error(f"Server error ({response.status}): {response.url}")
        else:
            self.logger.warning(f"HTTP error ({response.status}): {response.url}")

    def _handle_error(self, failure):
        """Handle request failures."""
        request = failure.request
        error_info = {
            "url": request.url,
            "error_type": "request_failure",
            "timestamp": datetime.now().isoformat(),
        }

        if failure.check(HttpError):
            response = failure.value.response
            error_info["status"] = response.status
        elif failure.check(DNSLookupError):
            error_info["error_message"] = "DNS lookup failed"
        elif failure.check(TimeoutError, TCPTimedOutError):
            error_info["error_message"] = "Request timed out"
        else:
            error_info["error_message"] = str(failure.value)

        self.errors.append(error_info)
        self.logger.error(f"Error on {request.url}: {error_info}")


# =============================================================================
# OUTPUT HANDLING
# =============================================================================


def generate_output_path(domain, scrape_type):
    """Generate output path with folder structure and timestamp."""
    # Get the script's directory as base
    script_dir = Path(__file__).parent
    scraped_sites_dir = script_dir / "scraped-sites"

    # Clean domain for folder name
    clean_domain = domain.replace("www.", "").replace(".", "-")

    # Create domain-specific folder
    domain_folder = scraped_sites_dir / clean_domain
    domain_folder.mkdir(parents=True, exist_ok=True)

    # Add timestamp to filename
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"{clean_domain}-{scrape_type}-{timestamp}"

    return domain_folder / filename


def save_results(spider, output_base):
    """Save crawl results to CSV files."""
    page_file = f"{output_base}.csv"
    unique_file = f"{output_base}_unique.csv"

    # Save page-level results with one image per row
    if spider.results:
        fieldnames = ["page_url", "page_title", "image_url", "image_alt"]
        with open(page_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for result in spider.results:
                images = json.loads(result["images"])
                for i, img in enumerate(images):
                    writer.writerow({
                        "page_url": result["page_url"] if i == 0 else "",
                        "page_title": result["page_title"] if i == 0 else "",
                        "image_url": img["src"],
                        "image_alt": img.get("alt", ""),
                    })

    # Save unique images
    if spider.all_images:
        with open(unique_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f, fieldnames=["src", "alt", "pages_found_on", "page_count"]
            )
            writer.writeheader()
            for img_url, img_data in sorted(spider.all_images.items()):
                writer.writerow({
                    "src": img_url,
                    "alt": img_data["alt"],
                    "pages_found_on": json.dumps(img_data["found_on"]),
                    "page_count": len(img_data["found_on"]),
                })

    # Print summary
    print("\n" + "=" * 60)
    print("SCRAPE COMPLETE")
    print("=" * 60)
    print(f"Domain:            {spider.config['domain']}")
    print(f"Scrape type:       {spider.config['scrape_type']}")
    print(f"Pages crawled:     {spider.pages_crawled}")
    print(f"Pages with images: {len(spider.results)}")
    print(f"Unique images:     {len(spider.all_images)}")
    print(f"Errors:            {len(spider.errors)}")
    print("-" * 60)
    print(f"Output files:")
    print(f"  Page results:    {page_file}")
    print(f"  Unique images:   {unique_file}")
    print("=" * 60)

    if spider.errors:
        print("\nErrors encountered (first 5):")
        for error in spider.errors[:5]:
            print(f"  - {error.get('url', 'Unknown')}: {error.get('status', error.get('error_message', 'Unknown'))}")


def save_results_excel(spider, output_base):
    """Save Full Monty results to an Excel workbook with separate sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        print("Falling back to CSV output.")
        save_results(spider, output_base)
        return

    xlsx_file = f"{output_base}.xlsx"
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Separate results by page type
    category_results = [r for r in spider.results if r.get("page_type") == "category"]
    product_results = [r for r in spider.results if r.get("page_type") == "product"]
    other_results = [r for r in spider.results if r.get("page_type") == "other"]

    # Column definitions: (Header, field_key, width)
    category_columns = [
        ("Page URL", "page_url", 60),
        ("Meta Title", "meta_title", 45),
        ("Meta Description", "meta_description", 55),
        ("H1", "h1", 40),
        ("Canonical URL", "canonical_url", 60),
        ("Word Count", "word_count", 12),
        ("Image Count", "total_image_count", 12),
        ("Internal Links", "internal_link_count", 14),
        ("External Links", "external_link_count", 14),
        ("OG Image", "og_image", 50),
        ("Has Schema", "has_schema_markup", 12),
        ("Schema Types", "schema_types", 30),
    ]

    product_columns = [
        ("Page URL", "page_url", 60),
        ("Product Name", "product_name", 40),
        ("Price", "product_price", 12),
        ("Currency", "product_currency", 10),
        ("SKU", "product_sku", 18),
        ("Brand", "product_brand", 20),
        ("Availability", "product_availability", 18),
        ("Product Description", "product_description", 55),
        ("Meta Title", "meta_title", 45),
        ("Meta Description", "meta_description", 55),
        ("H1", "h1", 40),
        ("Canonical URL", "canonical_url", 60),
        ("Word Count", "word_count", 12),
        ("Image Count", "total_image_count", 12),
        ("Internal Links", "internal_link_count", 14),
        ("External Links", "external_link_count", 14),
        ("OG Image", "og_image", 50),
        ("Has Schema", "has_schema_markup", 12),
        ("Schema Types", "schema_types", 30),
    ]

    def write_sheet(ws, columns, data, sheet_title):
        """Write data to a worksheet with formatting."""
        ws.title = sheet_title

        # Write headers
        for col_idx, (header, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write data rows
        for row_idx, result in enumerate(data, 2):
            for col_idx, (_, field_key, _) in enumerate(columns, 1):
                value = result.get(field_key, "")
                if isinstance(value, bool):
                    value = "Yes" if value else "No"
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Freeze header row and add auto-filter
        ws.freeze_panes = "A2"
        if data:
            ws.auto_filter.ref = ws.dimensions

    # Sheet 1: Categories (use default sheet)
    ws_categories = wb.active
    write_sheet(ws_categories, category_columns, category_results, "Categories")

    # Sheet 2: Products
    ws_products = wb.create_sheet()
    write_sheet(ws_products, product_columns, product_results, "Products")

    # Sheet 3: Other Pages (if any)
    if other_results:
        ws_other = wb.create_sheet()
        write_sheet(ws_other, category_columns, other_results, "Other Pages")

    # Sheet: Summary
    ws_summary = wb.create_sheet("Summary")
    summary_header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_data = [
        ("Domain", spider.config["domain"]),
        ("Scrape Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Pages Crawled", spider.pages_crawled),
        ("Category Pages", len(category_results)),
        ("Product Pages", len(product_results)),
        ("Other Pages", len(other_results)),
        ("Unique Images", len(spider.all_images)),
        ("Errors", len(spider.errors)),
    ]
    # Header row
    for col_idx, header in enumerate(["Metric", "Value"], 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = summary_header_font
        cell.fill = summary_header_fill
        cell.border = thin_border
    # Data rows
    for row_idx, (metric, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=1, value=metric).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=value).border = thin_border
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40

    wb.save(xlsx_file)

    # Print summary
    print("\n" + "=" * 60)
    print("FULL MONTY SCRAPE COMPLETE")
    print("=" * 60)
    print(f"Domain:            {spider.config['domain']}")
    print(f"Pages crawled:     {spider.pages_crawled}")
    print(f"Category pages:    {len(category_results)}")
    print(f"Product pages:     {len(product_results)}")
    print(f"Other pages:       {len(other_results)}")
    print(f"Unique images:     {len(spider.all_images)}")
    print(f"Errors:            {len(spider.errors)}")
    print("-" * 60)
    print(f"Output file:")
    print(f"  Excel report:    {xlsx_file}")
    print("=" * 60)

    if spider.errors:
        print("\nErrors encountered (first 5):")
        for error in spider.errors[:5]:
            print(f"  - {error.get('url', 'Unknown')}: {error.get('status', error.get('error_message', 'Unknown'))}")


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================


def run_scraper(domain, scrape_type, max_pages=0, delay=1.0, concurrent=2,
                seed_urls=None, known_product_urls=None, known_category_urls=None):
    """
    Run the scraper with given parameters.

    Args:
        domain: Website domain (e.g., "www.example.com")
        scrape_type: Type of pages to scrape ("category", "product", "blog", "all", "fullmonty")
        max_pages: Maximum pages to crawl (0 = unlimited)
        delay: Delay between requests in seconds
        concurrent: Number of concurrent requests
        seed_urls: Optional list of additional URLs to use as start pages
        known_product_urls: Optional list of URLs known to be product pages (for classification)
        known_category_urls: Optional list of URLs known to be category pages (for classification)

    Returns:
        Tuple of (output_base_filename, pages_crawled, unique_images_count)
    """
    if scrape_type not in SCRAPE_CONFIGS:
        print(f"Error: Invalid scrape type '{scrape_type}'")
        print(f"Valid types: {', '.join(SCRAPE_CONFIGS.keys())}")
        sys.exit(1)

    # Prepare config
    config = {
        "domain": domain,
        "scrape_type": scrape_type,
        "max_pages": max_pages,
        "delay": delay,
        "concurrent": concurrent,
        "min_image_width": 50,
        "min_image_height": 50,
        "seed_urls": seed_urls or [],
        "known_product_urls": known_product_urls or [],
        "known_category_urls": known_category_urls or [],
    }

    output_base = generate_output_path(domain, scrape_type)

    print("=" * 60)
    print("Website Scraper")
    print("=" * 60)
    print(f"Domain:            {domain}")
    print(f"Scrape type:       {scrape_type} ({SCRAPE_CONFIGS[scrape_type]['description']})")
    print(f"Download delay:    {delay}s")
    print(f"Concurrent:        {concurrent}")
    if max_pages > 0:
        print(f"Max pages:         {max_pages}")
    if config["seed_urls"]:
        print(f"Seed URLs:         {len(config['seed_urls'])} additional start pages")
    if config["known_product_urls"]:
        print(f"Known products:    {len(config['known_product_urls'])} product URLs for classification")
    if config["known_category_urls"]:
        print(f"Known categories:  {len(config['known_category_urls'])} category URLs for classification")
    print(f"Output folder:     {output_base.parent}")
    print(f"Output prefix:     {output_base.name}")
    print("=" * 60 + "\n")

    # Create and run crawler
    process = CrawlerProcess(settings=ScraperSpider.create_settings(config))

    # Create spider instance
    spider = ScraperSpider(config=config)

    # Create crawler
    crawler = process.create_crawler(ScraperSpider)

    # Connect save handler
    def on_spider_closed(spider):
        if spider.config["scrape_type"] == "fullmonty":
            save_results_excel(spider, output_base)
        else:
            save_results(spider, output_base)

    crawler.signals.connect(on_spider_closed, signal=signals.spider_closed)

    # Run the crawl
    process.crawl(crawler, config=config)
    process.start()

    return output_base


def main():
    """Parse arguments and run scraper."""
    parser = argparse.ArgumentParser(
        description="Universal Website Scraper - Extract images, product data, and SEO data from websites",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scraper.py --url "www.example.com" --type category
  python scraper.py --url "www.example.com" --type product --max-pages 100
  python scraper.py --url "www.example.com" --type all --delay 0.5
  python scraper.py --url "www.example.com" --type fullmonty

Scrape types:
  category  - Category/collection pages (/category/, /shop/, /collections/)
  product   - Product detail pages (/product/, /p/, /item/)
  blog      - Blog/news pages (/blog/, /news/, /articles/)
  all       - Entire website (follows all internal links)
  fullmonty - Full audit: categories + products with SEO, product & content data (Excel output)
        """
    )

    parser.add_argument(
        "--url", "-u",
        required=True,
        help="Website domain to scrape (e.g., www.example.com)"
    )
    parser.add_argument(
        "--type", "-t",
        required=True,
        choices=list(SCRAPE_CONFIGS.keys()),
        help="Type of pages to scrape"
    )
    parser.add_argument(
        "--max-pages", "-m",
        type=int,
        default=0,
        help="Maximum pages to crawl (0 = unlimited)"
    )
    parser.add_argument(
        "--delay", "-d",
        type=float,
        default=1.0,
        help="Delay between requests in seconds (default: 1.0)"
    )
    parser.add_argument(
        "--concurrent", "-c",
        type=int,
        default=2,
        help="Number of concurrent requests (default: 2)"
    )
    parser.add_argument(
        "--urls",
        type=str,
        default=None,
        help="Path to a text file containing seed URLs (one per line) to use as additional start pages"
    )
    parser.add_argument(
        "--product-urls",
        type=str,
        default=None,
        help="Path to a text file containing known product URLs (one per line) for page classification"
    )
    parser.add_argument(
        "--category-urls",
        type=str,
        default=None,
        help="Path to a text file containing known category URLs (one per line) for page classification"
    )

    args = parser.parse_args()

    # Clean up URL (remove protocol if provided)
    domain = args.url.replace("https://", "").replace("http://", "").rstrip("/")

    # Load seed URLs from file if provided
    seed_urls = []
    if args.urls:
        try:
            with open(args.urls, "r") as f:
                seed_urls = [line.strip() for line in f if line.strip()]
            print(f"Loaded {len(seed_urls)} seed URLs from {args.urls}")
        except FileNotFoundError:
            print(f"Warning: URL file not found: {args.urls}")

    # Load known product URLs for classification
    known_product_urls = []
    if args.product_urls:
        try:
            with open(args.product_urls, "r") as f:
                known_product_urls = [line.strip() for line in f if line.strip()]
            print(f"Loaded {len(known_product_urls)} known product URLs from {args.product_urls}")
        except FileNotFoundError:
            print(f"Warning: Product URL file not found: {args.product_urls}")

    # Load known category URLs for classification
    known_category_urls = []
    if args.category_urls:
        try:
            with open(args.category_urls, "r") as f:
                known_category_urls = [line.strip() for line in f if line.strip()]
            print(f"Loaded {len(known_category_urls)} known category URLs from {args.category_urls}")
        except FileNotFoundError:
            print(f"Warning: Category URL file not found: {args.category_urls}")

    run_scraper(
        domain=domain,
        scrape_type=args.type,
        max_pages=args.max_pages,
        delay=args.delay,
        concurrent=args.concurrent,
        seed_urls=seed_urls,
        known_product_urls=known_product_urls,
        known_category_urls=known_category_urls,
    )


if __name__ == "__main__":
    main()
